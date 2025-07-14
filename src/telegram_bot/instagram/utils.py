import re
import os
from datetime import datetime
from datetime import timedelta
import pandas as pd
import logging

import openpyxl
from openpyxl.styles import Alignment

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def format_excel_file(filepath: str) -> str:
    """ Apply formatting to an Excel file

    Args:
        filepath: Path to the Excel file

    Returns:
        filepath: Path to the formatted Excel file

    """
    # Load the Excel file to apply formatting
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Set the alignment to center for all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column-1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                logger.error(f"Error formatting Excel file: {e}")
        adjusted_width = (max_length + 2)  # Add extra padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(filepath)
    return filepath


# Utility: Cleanup old files
def cleanup_files(user_dir: str, retention_period: int = 2):
    now = datetime.now()
    for root, _, files in os.walk(user_dir):
        for file in files:
            file_path = os.path.join(root, file)
            file_creation_time = datetime.fromtimestamp(os.path.getctime(file_path))
            if now - file_creation_time > timedelta(days=retention_period):
                os.remove(file_path)
                logger.info(f"Deleted old file: {file_path}")



def sanitize_instagram_input(user_input: str) -> str:
    user_input = user_input.replace("#", "").replace("@", "")
    if "instagram.com" in user_input:
        match = re.search(r"(?:https?://)?(?:www\.)?instagram\.com/([A-Za-z0-9_.]+)", user_input)
        if match:
            user_input = match.group(1)
    return user_input


def create_resource(user_id: int, name: str, data_list: list[dict]) -> str:
    # Create user directory
    user_dir = f"./tmp/{user_id}"
    os.makedirs(user_dir, exist_ok=True)
    cleanup_files(user_dir)

    # Sanitize name for filename
    sanitized_name = re.sub(r'[\/:*?"<>| ]', '_', name)[:15]

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"{timestamp}_{sanitized_name}_ig.xlsx"
    filepath = os.path.join(user_dir, filename)

    # Create and save Excel file
    df = pd.DataFrame(data_list)
    df.to_excel(filepath, index=False)
    format_excel_file(filepath)

    return filename
